import openpyxl
from tkinter import Tk, messagebox, filedialog
import tkinter as tk
from openpyxl import Workbook
from openpyxl.styles import Font
from PyPDF2 import PdfReader
from tkinter.filedialog import askopenfilename
import matplotlib.pyplot as plt
import os
import tempfile
import win32com.client as win32

clientes_file = "clientes.xlsx"

def carregar_clientes():
    try:
        workbook = openpyxl.load_workbook(clientes_file)
        sheet = workbook.active
        clientes = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            nome, cpf, tipo = row
            clientes[nome] = tipo
        workbook.close()
        return clientes
    except FileNotFoundError:
        return {}

def salvar_cliente(nome, cpf, tipo):
    global clientes_file

    # Continuar com a função normal de salvar o cliente no arquivo
    try:
        workbook = openpyxl.load_workbook(clientes_file)
    except FileNotFoundError:
        # Se o arquivo não existir, criar um novo arquivo com o formato esperado
        workbook = Workbook()
        workbook.active.title = "Clientes"
        sheet = workbook.active
        sheet.append(["Nome", "CPF", "Tipo"])
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        # Salvar o arquivo no diretório atual
        workbook.save(clientes_file)

    sheet = workbook.active
    nomes_existentes = [cell.value for cell in sheet['A'][1:]]
    if nome in nomes_existentes:
        print(f"O cliente '{nome}' já está cadastrado.")
        return

    sheet.append([nome, cpf, tipo])
    workbook.save(clientes_file)

def excluir_cliente(nome_cliente):
    clientes = carregar_clientes()
    if clientes:
        workbook = openpyxl.load_workbook(clientes_file)
        sheet = workbook.active

        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == nome_cliente:
                sheet.delete_rows(row)
                workbook.save(clientes_file)
                messagebox.showinfo("Sucesso", f"Cliente '{
                                    nome_cliente}' excluído com sucesso!")
                abrir_janela_digitar_nome_cliente()
                return

        messagebox.showwarning(
            "Aviso", f"Não foi encontrado um cliente com o nome '{nome_cliente}'.")
    else:
        messagebox.showwarning("Aviso", "Não há clientes para excluir.")


def adicionar_cliente():
    nome = entry_nome.get()
    tipo = tipo_var.get()
    cpf = entry_cpf.get()

    salvar_cliente(nome, cpf, tipo)

    messagebox.showinfo("Sucesso", "Cliente adicionado com sucesso!")
    nova_janela.destroy()
    abrir_janela_digitar_nome_cliente()


def abrir_nova_janela():
    global nova_janela, entry_nome, entry_cpf, tipo_var
    fechar_janelas()
    nova_janela = tk.Toplevel(root)
    nova_janela.title("Adicionar Cliente")

    label_nome = tk.Label(nova_janela, text="Nome:", font=("Helvetica", 12))
    label_nome.grid(row=0, column=0, padx=10, pady=5, sticky="w")
    entry_nome = tk.Entry(nova_janela, font=("Helvetica", 12))
    entry_nome.grid(row=0, column=1, padx=10, pady=5)

    label_cpf = tk.Label(nova_janela, text="CPF:", font=("Helvetica", 12))
    label_cpf.grid(row=1, column=0, padx=10, pady=5, sticky="w")
    entry_cpf = tk.Entry(nova_janela, font=("Helvetica", 12))
    entry_cpf.grid(row=1, column=1, padx=10, pady=5)

    label_tipo = tk.Label(
        nova_janela, text="Selecione o tipo:", font=("Helvetica", 12))
    label_tipo.grid(row=2, column=0, padx=10, pady=5, sticky="w")

    tipos = ["Arrojado", "Conservador", "Moderado", "Agressivo", "Balanceado"]
    tipo_var = tk.StringVar(nova_janela)
    tipo_var.set(tipos[0])

    option_menu = tk.OptionMenu(nova_janela, tipo_var, *tipos)
    option_menu.grid(row=2, column=1, padx=10, pady=5, sticky="ew")

    botao_adicionar = tk.Button(
        nova_janela, text="Adicionar", width=10, command=adicionar_cliente, font=("Helvetica", 12))
    botao_adicionar.grid(row=3, column=0, columnspan=2, padx=10, pady=10)

    botao_excluir = tk.Button(
        nova_janela, text="Excluir", width=10, command=lambda: excluir_cliente(entry_nome.get()), font=("Helvetica", 12))
    botao_excluir.grid(row=4, column=0, columnspan=2, padx=10, pady=10)


def abrir_janela_digitar_nome_cliente():
    fechar_janelas()
    global janela_digitar_nome_cliente
    janela_digitar_nome_cliente = tk.Toplevel(root)
    janela_digitar_nome_cliente.title("Digite o Nome do Cliente")

    frame = tk.Frame(janela_digitar_nome_cliente)
    frame.pack(pady=10)

    label_nome = tk.Label(frame, text="Nome do Cliente:",
                          font=("Helvetica", 12))
    label_nome.grid(row=0, column=0, padx=10, pady=5, sticky="w")

    entry_nome_cliente = tk.Entry(frame, font=("Helvetica", 12))
    entry_nome_cliente.grid(row=0, column=1, padx=10, pady=5)

    botao_ok = tk.Button(frame, text="Ok", command=lambda: verificar_tipo_cliente(
        entry_nome_cliente.get(), janela_digitar_nome_cliente), font=("Helvetica", 12))
    botao_ok.grid(row=1, column=0, columnspan=2, pady=10)

    botao_fechar = tk.Button(
        frame, text="Fechar", command=janela_digitar_nome_cliente.destroy, font=("Helvetica", 12))
    botao_fechar.grid(row=2, column=0, columnspan=2, pady=10)

    # Retorna o nome do cliente digitado
    return entry_nome_cliente.get()


def fechar_janelas_secundarias():
    for widget in root.winfo_children():
        if isinstance(widget, tk.Toplevel) and widget != janela_digitar_nome_cliente:
            widget.destroy()


def abrir_janela_selecionar_arquivo():
    fechar_janelas()
    janela_selecionar_arquivo = tk.Toplevel(root)
    janela_selecionar_arquivo.title("Selecionar Arquivo Excel")

    label = tk.Label(janela_selecionar_arquivo,
                     text="SELECIONE UM ARQUIVO EXCEL", font=("Helvetica", 12))
    label.pack(pady=10)

    botao_ok = tk.Button(janela_selecionar_arquivo, text="Ok",
                         command=selecionar_arquivo, font=("Helvetica", 12))
    botao_ok.pack()


def nao():
    fechar_janelas_secundarias()
    abrir_janela_digitar_nome_cliente()


def sim():
    abrir_nova_janela()


def fechar_janelas():
    for widget in root.winfo_children():
        if isinstance(widget, tk.Toplevel):
            widget.destroy()


def selecionar_arquivo():
    arquivo_excel = filedialog.askopenfilename(
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")])
    if arquivo_excel:
        carregar_dados(arquivo_excel)


def carregar_dados(arquivo_excel):
    try:
        workbook = openpyxl.load_workbook(arquivo_excel)
        print("Arquivo carregado com sucesso:", arquivo_excel)
        abrir_janela_selecionar_pdf()
    except Exception as e:
        print("Erro ao carregar o arquivo:", e)


def abrir_janela_selecionar_pdf():
    fechar_janelas()
    janela_selecionar_pdf = tk.Toplevel(root)
    janela_selecionar_pdf.title("Selecionar Arquivo PDF")

    label = tk.Label(janela_selecionar_pdf,
                     text="SELECIONE UM ARQUIVO PDF", font=("Helvetica", 12))
    label.pack(pady=10)

    botao_ok = tk.Button(janela_selecionar_pdf, text="Ok",
                         command=selecionar_pdf, font=("Helvetica", 12))
    botao_ok.pack()


def selecionar_pdf():
    arquivo_pdf = filedialog.askopenfilename(
        filetypes=[("Arquivos PDF", "*.pdf"), ("Todos os arquivos", "*.*")])
    if arquivo_pdf:
        print("Arquivo PDF selecionado:", arquivo_pdf)
        # Chamada para executar o segundo código com base no arquivo PDF selecionado
        executar_segundo_codigo(arquivo_pdf)


def verificar_tipo_cliente(nome_cliente, janela):
    global nome_digitado
    # Convertendo o nome digitado para minúsculas
    nome_digitado = nome_cliente.lower()
    clientes = carregar_clientes()
    # Convertendo os nomes existentes para minúsculas
    nome_cliente_minusculo = nome_cliente.lower()

    if nome_cliente_minusculo in clientes:
        tipo_cliente = clientes[nome_cliente_minusculo]
        print(f"O tipo do cliente '{nome_cliente}' é '{tipo_cliente}'.")
        # Chama a função para executar o segundo código com base no tipo do cliente
        # Corrigido para passar o tipo do cliente
        executar_segundo_codigo(tipo_cliente)
    else:
        print(f"Não foi encontrado um cliente com o nome '{nome_cliente}'.")
    janela.destroy()


def executar_segundo_codigo(tipo_cliente):
    arquivo_excel, arquivo_pdf = selecionar_arquivos()
    if arquivo_excel and arquivo_pdf:
        # O tipo do cliente é usado como a palavra a ser pesquisada no arquivo Excel
        palavra = tipo_cliente
        resultados_excel = encontrar_palavra(arquivo_excel, palavra)
        if resultados_excel:
            for ws, cell in resultados_excel:
                print(f'A palavra "{palavra}" foi encontrada na planilha "{
                      ws.title}" na célula {cell.coordinate} com valor "{cell.value}"')
                imprimir_linhas(ws, cell.row, cell.column, palavra)
        else:
            print(f'A palavra "{
                  palavra}" não foi encontrada no arquivo Excel.')
        # Gerar o gráfico com base no tipo do cliente antes de chamar a função 'convergencia'
        unificar_graficos(tipo_cliente, arquivo_pdf)
        convergencia(arquivo_pdf, tipo_cliente)
        # Fechar todas as janelas de plotagem
        plt.close()
        # Sair do loop principal, impedindo a exibição de outras janelas gráficas
        return
    else:
        print("Nenhum arquivo selecionado. Encerrando...")


def selecionar_arquivo_excel():
    arquivo_excel = filedialog.askopenfilename(
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")])
    return arquivo_excel


def convergencia(arquivo_pdf, tipo_cliente):
    num_paginas = len(PdfReader(arquivo_pdf).pages)
    for pagina in range(num_paginas):
        arquivo = (
            PdfReader(arquivo_pdf).pages[pagina]).extract_text().splitlines()
        if "Consolidada |Alocação por Estratégia" in arquivo[0][:36]:
            pagina_procurada = pagina
            break
    arquivo = (
        PdfReader(arquivo_pdf).pages[pagina_procurada]).extract_text().splitlines()
    categoria_individual = []
    categoria_geral = []
    cont = 0
    for elemento in arquivo:
        if len(elemento) > 56 and cont > 0:
            categoria = elemento[:-53]
            categoria = " ".join(categoria[:-6].split()).split(" - ")[0]
            if categoria == "Fundos de Ações" or categoria == "Ações":
                categoria = "Renda Variável"
            elif categoria == "FIIs":
                categoria = "Alternativos"
            porcentagem = elemento[:-53]
            if elemento.count("-") > 0:
                porcentagem = porcentagem.replace(" - ", " ")
            porcentagem = porcentagem.split()[-1]
            try:
                porcentagem = float(porcentagem.replace(
                    ",", ".").replace("%", ""))
            except ValueError:
                porcentagem = 0.0
            categoria_individual.append(categoria)
            categoria_individual.append(porcentagem)
            categoria_geral.append(categoria_individual)
            break
        elif len(elemento) <= 56 and cont > 0:
            categoria = " ".join(elemento[:-6].split()).split(" - ")[0]
            if categoria == "Fundos de Ações" or categoria == "Ações":
                categoria = "Renda Variável"
            elif categoria == "FIIs":
                categoria = "Alternativos"
            categoria_individual.append(categoria)
            porcentagem = elemento[-6:].strip()
            try:
                porcentagem = float(porcentagem.replace(
                    ",", ".").replace("%", ""))
            except ValueError:
                porcentagem = 0.0
            categoria_individual.append(porcentagem)
            categoria_geral.append(categoria_individual)
        categoria_individual = []
        cont += 1

    soma = {}
    for sublista in categoria_geral:
        if sublista[0] in soma:
            soma[sublista[0]] += sublista[1]
        else:
            soma[sublista[0]] = sublista[1]
    lista_categorias_soma = [[k, v] for k, v in soma.items()]

    # Retorna a lista de categorias com as somas das porcentagens
    return lista_categorias_soma


def enviar_email(cliente, perfil, texto, path_combined_image):
    outlook = win32.Dispatch('outlook.application')
    email = outlook.CreateItem(0)
    email.To = 'pedro.moreira@garzaif.com.br'
    email.Subject = f"Convergência - {cliente}"
    email.HTMLBody = texto + "<img src='cid:imagem_cid'>"
    attachment = email.Attachments.Add(path_combined_image)
    attachment.PropertyAccessor.SetProperty(
        "http://schemas.microsoft.com/mapi/proptag/0x3712001E", "imagem_cid")
    email.Send()
    outlook.Quit()


def ajustar_categorias_pdf(categorias_pdf):
    ajustes = {
        "Renda Fixa Pós Fixado": "Pos Fixado",
        "Renda Fixa Pré Fixado": "Pre Fixado",
        "Renda Fixa Inflação": "Inflação",
        "Renda Fixa Global": "RF Global",
        "Fundos Multimercado": "Multimercado",
        "Renda Variável Global": "Renda Variável Global",
        "FIIs": "Alternativos",
        "Alternativos": "Alternativos",
        "Ações": "Renda Variável",
        "Fundos de Ações": "Renda Variável"
    }
    return [ajustes.get(categoria, categoria) for categoria in categorias_pdf]


def unificar_graficos(tipo_cliente, arquivo_pdf):
    # Dados para o gráfico de alocação por categoria do Excel
    categorias_excel = ["Pos Fixado", "Pre Fixado", "Inflação", "RF Global",
                        "Multimercado", "Renda Variável", "Renda Variável Global", "Alternativos", ]
    porcentagens_excel = None

    if tipo_cliente == "Balanceado":
        porcentagens_excel = [46.85, 7.5, 22.1, 0, 8.04, 9.6, 6, 0]
    elif tipo_cliente == "Moderado":
        porcentagens_excel = [62.05, 5, 20.23, 0, 4.32, 4.8, 3.6, 0]
    elif tipo_cliente == "Conservador":
        porcentagens_excel = [88.5, 2.5, 5.8, 0, 3.2, 0, 0, 0]
    elif tipo_cliente == "Arrojado":
        porcentagens_excel = [34.66, 8.75, 22.27, 0, 9.02, 13.3, 9.5, 2.5]
    else:
        porcentagens_excel = [22.48, 10, 22.52, 0, 10, 17, 13, 5]

    # Obter as somas das porcentagens de alocação para cada categoria do PDF
    lista_categorias_soma = convergencia(arquivo_pdf, tipo_cliente)

    # Extrair categorias e somas das porcentagens do resultado da função convergencia
    categorias_pdf = [categoria[0] for categoria in lista_categorias_soma]
    porcentagens_soma_pdf = [porcentagem[1]
                             for porcentagem in lista_categorias_soma]

    # Ajustar categorias do PDF para corresponder às categorias do Excel
    categorias_pdf = ajustar_categorias_pdf(categorias_pdf)

    # Unificar dados
    categorias_unificadas = categorias_excel + \
        list(set(categorias_pdf) - set(categorias_excel))
    porcentagens_unificadas_excel = [porcentagens_excel[categorias_excel.index(
        categoria)] if categoria in categorias_excel else 0 for categoria in categorias_unificadas]
    porcentagens_unificadas_pdf = [porcentagens_soma_pdf[categorias_pdf.index(
        categoria)] if categoria in categorias_pdf else 0 for categoria in categorias_unificadas]

    # Garantir que ambas as listas tenham o mesmo comprimento preenchendo com zero
    max_length = max(len(porcentagens_unificadas_excel),
                     len(porcentagens_unificadas_pdf))
    porcentagens_unificadas_excel += [0] * \
        (max_length - len(porcentagens_unificadas_excel))
    porcentagens_unificadas_pdf += [0] * \
        (max_length - len(porcentagens_unificadas_pdf))

    # Calcular a diferença entre as porcentagens do Excel e do PDF
    diferenca_porcentagens = [pdf - excel for pdf, excel in zip(
        porcentagens_unificadas_pdf, porcentagens_unificadas_excel)]

    # Tomar os valores absolutos da diferença
    diferenca_porcentagens_abs = [abs(valor)
                                  for valor in diferenca_porcentagens]

    # Plotar gráfico unificado com as diferenças absolutas
    plt.figure(figsize=(12, 6))
    bar_width = 0.35
    index = range(len(categorias_unificadas))

    # Barras para as porcentagens do Excel
    plt.bar(index, porcentagens_unificadas_excel, bar_width,
            label='Recomendado', color='skyblue', edgecolor='black')

    # Barras para as porcentagens do PDF
    plt.bar([i + bar_width for i in index], porcentagens_unificadas_pdf,
            bar_width, label='Atual', color='lightgreen', edgecolor='black')

    # Barras para as diferenças absolutas
    plt.bar([i + bar_width / 2 for i in index], diferenca_porcentagens_abs,
            bar_width, label='Diferença Absoluta', color='orange', edgecolor='black')

    # Adicionar rótulos para as barras que exibem a diferença absoluta
    for i in index:
        if diferenca_porcentagens[i] >= 0:
            plt.text(i + bar_width / 2, diferenca_porcentagens_abs[i] + 0.5, f'+{
                     round(diferenca_porcentagens[i], 2)}%', fontsize=8, color='black', ha='center')
        else:
            plt.text(i + bar_width / 2, diferenca_porcentagens_abs[i] + 0.5, f'{
                     round(diferenca_porcentagens[i], 2)}%', fontsize=8, color='black', ha='center')

        # Adicionar valores acima das colunas "Atual" e "Recomendado" com apenas duas casas decimais
        plt.text(i, porcentagens_unificadas_excel[i] + 1, f"{
                 porcentagens_unificadas_excel[i]:.2f}%", ha='center', va='bottom', color='black')
        plt.text(i + bar_width, porcentagens_unificadas_pdf[i] + 1, f"{
                 porcentagens_unificadas_pdf[i]:.2f}%", ha='center', va='bottom', color='black')

    # Personalizar o gráfico
    plt.xlabel('Categorias', fontsize=12)
    plt.ylabel('Porcentagem de alocação', fontsize=12)
    plt.title(
        'Alocação por Categoria - Comparação entre PDF e Excel (Diferença Absoluta)', fontsize=14)
    plt.xticks([i + bar_width / 2 for i in index],
               categorias_unificadas, rotation=45, fontsize=10)
    plt.yticks(fontsize=10)
    plt.legend({'Recomendado': 'blue', 'Atual': 'orange',
               'Diferença': 'green'}, fontsize=10)
    # Salvar o gráfico como uma imagem temporária
    temp_image_file = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    plt.savefig(temp_image_file.name)
    plt.close()

    # Enviar o e-mail com o gráfico como anexo
    enviar_email(tipo_cliente, "Gráfico de Alocação por Categoria",
                 "Segue o gráfico de alocação por categoria:", temp_image_file.name)

    # Exibir o gráfico


def ajustar_categorias_pdf(categorias_pdf):
    ajustes = {
        "Renda Fixa Pós Fixado": "Pos Fixado",
        "Renda Fixa Pré Fixado": "Pre Fixado",
        "Renda Fixa Inflação": "Inflação",
        "RF Global": "RF Global",
        "Fundos Multimercado": "Multimercado",
        "Renda Variável": "Renda Variável",
        "Renda Variável Global": "Renda Variável Global",
        "FIIs": "Alternativos",
        "Carteira Recomendada": "Renda Variável Global",
        "Alternativos": "Alternativos",
        "Ações": "Renda Variável",
        "Fundos de Ações": "Renda Variável"


    }
    return [ajustes[categoria] if categoria in ajustes else categoria for categoria in categorias_pdf]


def encontrar_palavra(arquivo, palavra):
    wb = openpyxl.load_workbook(arquivo)
    resultado = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if palavra.lower() in str(cell.value).lower():
                    resultado.append((ws, cell))

    return resultado


def imprimir_linhas(ws, row, col, palavra):
    print(f"\n{palavra}\tNova Tática")
    for i in range(row + 1, row + 9):
        for j in range(col, col + 6):
            cell = ws.cell(row=i, column=j)
            if j == col:
                print(cell.value, end="\t")
            else:
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    print(f"{cell.value:.4f}", end="\t")
                else:
                    print(cell.value, end="\t")
        print()


def selecionar_arquivos():
    Tk().withdraw()
    filename_excel = askopenfilename(
        filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")])

    root = Tk()
    root.withdraw()
    filename_pdf = filedialog.askopenfilename(
        filetypes=[("Arquivos PDF", "*.pdf"), ("Todos os arquivos", "*.*")])
    return filename_excel, filename_pdf


root = tk.Tk()
root.title("Análise de Investimentos")
root.geometry("400x200")

label = tk.Label(root, text="Deseja cadastrar um cliente?",
                 font=("Helvetica", 12))
label.pack(pady=10)

frame = tk.Frame(root)
frame.pack(pady=10)

sim_button = tk.Button(frame, text="Sim", width=10,
                       command=sim, font=("Helvetica", 12))
sim_button.grid(row=0, column=0, padx=10)

nao_button = tk.Button(frame, text="Não", width=10,
                       command=nao, font=("Helvetica", 12))
nao_button.grid(row=0, column=1, padx=10)

root.mainloop()
